"""
report_engine.py — Sigma Cement monthly PDF report generator
=============================================================
Deterministic (no AI) engine: openpyxl extraction -> analytics -> matplotlib charts
-> reportlab PDF, reproducing the management report format (Exec Summary, Downtime
Study, Power, Product sections, Recipe Analysis, Stock+Pool policy, Silos,
Insights, Recommendations).

House rules encoded here (do not change without Muath's approval):
- Summary sheet is the AUTHORITATIVE source for monthly totals.
- All clinker variants consolidated as "Total Clinker".
- M50 ships 100% in bulk — no packing reconciliation for it.
- Grey pool (Clinker J + M) feeds M50: >= 6 months stock policy.
- White pool (Clinker ALB + SFW + RAK) feeds all whites: >= 4 months policy.
- Naming maps: daily "Super white Special" == M10 ; Power register "Power white-R" == CEM I 52.5R.
- Zero-production days flagged; Fridays are the weekly planned-stop regime.
"""
import io, os, json, calendar, tempfile
from datetime import datetime
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook

os.environ.setdefault('MPLCONFIGDIR', '/tmp/mpl')  # writable font-cache dir on Railway

# NOTE: matplotlib / reportlab / PIL are imported LAZILY inside make_charts()
# and build_pdf() so that `import report_engine` at bot startup stays light and
# cannot crash the service even if a plotting dependency has install issues —
# any such problem surfaces as a clear /report error message instead.

# ------------------------------------------------------------------ CONFIG ---
NAME_MAP = {'Super white Special': 'M10', 'Super white Spec': 'M10',
            'CEM I 52,5 R': 'CEM I 52.5R', 'Power white-R': 'CEM I 52.5R',
            'Wateen M10': 'M10', 'Wateen M50': 'M50'}

LIMITS = {  # tph_min, blaine_min, blaine_max, hex color
 'Power white':  (19.5, 3900, 4400, '#d35400'),
 'Super white':  (19.5, 4400, 5200, '#2e6db4'),
 'Eco white':    (19.5, 4800, 5700, '#27ae60'),
 'CEM I 52.5R':  (19.0, 3900, 4400, '#5d6d7e'),
 'M50':          (19.0, 3900, 4300, '#8e44ad'),
 'M10':          (19.0, 3800, 4500, '#16a085')}

MAT_ALIASES = {  # normalized (lower, no dots, single spaces) -> canonical
 'clinker roy': 'Clinker ROY', 'clinker r': 'Clinker ROY',
 'clinker sfw': 'Clinker SFW', 'clinker s': 'Clinker SFW',
 'clinker rak': 'Clinker RAK', 'clinker rk': 'Clinker RAK',
 'clinker j': 'Clinker J', 'clinker alb': 'Clinker ALB', 'clinker m': 'Clinker M',
 'limestone hg': 'Limestone HG', 'limestone lg': 'Limestone LG',
 'sand': 'Sand', 'sand & silica': 'Sand', 'sand& silica': 'Sand',
 'pozzolana': 'Pozzolana', 'gypsum': 'Gypsum', 'gypsum az': 'Gypsum AZ',
 'gypsum taf': 'Gypsum Taf', 'additive': 'Additive', 'grinding aids': 'Grinding aids',
 'quality enhancer': 'Quality enhancer', 'air entraining agent': 'Air entraining agent'}

def canon_mat(name):
    n = ' '.join(str(name).replace('.', '').split()).lower()
    return MAT_ALIASES.get(n, str(name).strip())

GREY_POOL  = ['Clinker J', 'Clinker M']            # M50 -> >= 6 months
WHITE_POOL = ['Clinker ALB', 'Clinker SFW', 'Clinker RAK']  # whites -> >= 4 months
GREY_FLOOR, WHITE_FLOOR = 6.0, 4.0
BULK_PRODUCTS = {'M50'}                            # no packing reconciliation

STOP_CATS = [  # (category, keyword tests on lowercase reason)
 ('Weekly holiday (Friday)',        lambda r,d: 'holiday' in r or 'week end' in r or 'weekend' in r),
 ('Silos full / dispatch',          lambda r,d: 'full silo' in r or 'full for level' in r or 'high level on silo' in r),
 ('Process fan (electrical)',       lambda r,d: 'process fan' in r or 'disconeccector' in r or 'disconnector' in r),
 ('Changeover / silo logistics',    lambda r,d: 'switching product' in r or 'convert to' in r or 'emtying silo' in r or 'emptying silo' in r or ('emptying' in r and 'switch' in r)),
 ('External power outage',          lambda r,d: 'power outage' in r or 'electricity supplier' in r),
 ('Human / operational error',      lambda r,d: 'mistake' in r or 'arrival delay' in r or 'samples failed' in r),
 ('Mechanical',                     lambda r,d: d == 'Mechanical'),
 ('Electrical (other)',             lambda r,d: d == 'Electrical'),
]

BLUE_HEX, DBLUE_HEX = '#1a5fa8', '#164e87'
LGREY_HEX, RED_HEX = '#f2f4f6', '#c0392b'
ORANGE_HEX, GREEN_HEX = '#e67e22', '#27ae60'

# ------------------------------------------------------------- EXTRACTION ---
def _f(v):
    try:
        x = float(v); return None if x != x else x
    except (TypeError, ValueError): return None

def extract(file_bytes):
    """Read the workbook into a plain-dict structure. Summary sheet = truth.
    Layout-adaptive: column positions come from header LABELS (supports both the
    2026 H1 template 'Daily report N' and the newer 'N June' template)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    D = {'products': {}, 'plant': {}, 'stock_close': {}, 'stock_open': {},
         'imports': {}, 'silos': {}, 'daily': {}, 'stoppages': [], 'pi': {},
         'missing_days': [], 'zero_days': []}

    sname = next((s for s in wb.sheetnames if 'summary' in s.lower()), None)
    if not sname:
        raise ValueError('Summary sheet not found — cannot build report (house rule).')
    rows = [list(r) for r in wb[sname].iter_rows(values_only=True)]

    # ---- product metric columns: mapped from the header row labels ----
    def norm(s): return ' '.join(str(s).replace('-', ' ').split()).lower()
    HDR_MAP = [('prod', 'production'), ('tph', 'mill productivity'),
               ('spc_mill', 'mill spc'), ('spc_plant', 'total spc'),
               ('hours', 'running hours'), ('ck', 'c/k'), ('clinker', 'clinker'),
               ('limestone', 'limestone'), ('gypsum', 'gypsum'), ('sand', 'sand'),
               ('pozzolana', 'pozzolana'), ('blaine', 'blaine'), ('r45', 'r45'),
               ('wi', 'whiteness')]
    colmap = {}
    hdr_row = rows[0]
    for ci, h in enumerate(hdr_row):
        if h is None: continue
        hn = norm(h)
        for key, lab in HDR_MAP:
            if key in colmap: continue
            if lab == 'production' and hn == 'production':   # first 'Production' = tons; second = '%'
                colmap[key] = ci
            elif lab != 'production' and lab in hn:
                colmap[key] = ci
    for r in rows[1:14]:
        p = str(r[0]).strip() if r[0] else ''
        p = NAME_MAP.get(p, p)
        if p in LIMITS and _f(r[colmap.get('prod', 1)]):
            D['products'][p] = {k: (_f(r[ci]) if ci is not None and ci < len(r) else None)
                                for k, ci in ((k, colmap.get(k)) for k, _ in HDR_MAP)}

    # ---- plant totals row: label-driven ----
    for i, r in enumerate(rows):
        if norm(r[0] or '').startswith('availability') and i + 1 < len(rows):
            labels = [norm(x or '') for x in r]
            vals = rows[i + 1]
            def by(*keys):
                for ci, lab in enumerate(labels):
                    if lab and all(k in lab for k in keys) and ci < len(vals):
                        return _f(vals[ci])
                return None
            def by_prefix(pfx):
                for ci, lab in enumerate(labels):
                    if lab.startswith(pfx) and ci < len(vals):
                        return _f(vals[ci])
                return None
            D['plant'] = {'availability': by('availability'), 'utilization': by('utilization'),
                          'prod': by('production'), 'avg_tph': by_prefix('av '),
                          'kwh': by('power'), 'spc': by('spc'),
                          'hours': by('running'), 'cost': by('electricity')}
            break

    # ---- final raw-material stock: anchored on the 'Raw material' banner ----
    for i, r in enumerate(rows):
        if r[0] and 'raw material' in str(r[0]).lower():
            names_row = next((rows[j] for j in range(i + 1, min(i + 3, len(rows)))
                              if any(c and 'clinker' in str(c).lower() for c in rows[j])), None)
            if names_row:
                j = rows.index(names_row)
                vals_row = rows[j + 1] if j + 1 < len(rows) else []
                for name, v in zip(names_row, vals_row):
                    if name and _f(v) is not None:
                        D['stock_close'][canon_mat(name)] = _f(v)
            break

    # ---- silo levels ----
    for r in rows:
        a = str(r[0]).strip().lower() if r[0] else ''
        if 'intal' in a or 'initial level' in a:
            D['silos']['start'] = [_f(x) for x in r[2:7]]
        elif a.startswith('end level'):
            vals = [_f(x) for x in r[2:7]]
            if any(v is not None for v in vals): D['silos']['end'] = vals
        elif a.startswith('paking') or a.startswith('packing'):
            D['silos']['packing_raw'] = [_f(x) for x in r[2:9]]

    # ---- Stock sheet: opening + imports (canonical names) ----
    if 'Stock' in wb.sheetnames:
        mode = None
        for r in wb['Stock'].iter_rows(values_only=True):
            a = str(r[0]).strip() if r[0] else ''
            if 'Initial' in a: mode = 'open'
            elif 'Daily import' in a: mode = 'imp'
            elif 'Material daily' in a: mode = None
            mat = canon_mat(r[2]) if r[2] else ''
            if mode and mat:
                nums = [v for v in (_f(x) for x in r[3:34]) if v is not None]
                if mode == 'open' and nums: D['stock_open'][mat] = nums[0]
                elif mode == 'imp': D['imports'][mat] = sum(nums)

    # ---- PI sheet ----
    if 'PI' in wb.sheetnames:
        pirows = [list(r) for r in wb['PI'].iter_rows(values_only=True)]
        dates = pirows[0][3:34]
        days = [d.day for d in dates if isinstance(d, datetime)]
        for lab_key, lab_txt in [('planned', 'Planned'), ('incident', 'Incident'), ('net', 'net operation')]:
            for r in pirows[1:]:
                if r[1] and lab_txt.lower() in str(r[1]).lower():
                    D['pi'][lab_key] = {d: _f(v) for d, v in zip(days, r[3:3 + len(days)])}
                    break

    # ---- Daily sheets ----
    import re
    daily = []
    for s in wb.sheetnames:
        m = re.match(r'Daily report\s+(\d+)', s, re.IGNORECASE) or re.match(r'(\d+)\s+\w+', s.strip())
        if m: daily.append((int(m.group(1)), s))
    daily.sort()
    have = [d for d, _ in daily]
    if have:
        D['missing_days'] = [d for d in range(1, max(have) + 1) if d not in have]
    METRIC_LABELS = {'prod': 'production', 'hours': 'running hours', 'tph': 'mill productiv',
                     'spc_mill': 'mill specific', 'spc_plant': 'total specific',
                     'ck': 'c/k', 'clinker': 'clinker', 'blaine': 'blaine',
                     'r45': 'r45', 'wi': 'whiteness'}
    for day, sheet in daily:
        rws = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        heads = [NAME_MAP.get(str(h).strip(), str(h).strip()) if h else '' for h in rws[0]]
        rowmap = {}
        for ri, r in enumerate(rws[:18]):
            lab = norm(r[0] or '')
            if not lab: continue
            for k, want in METRIC_LABELS.items():
                if k not in rowmap and lab.startswith(want.split()[0]) and want.split()[0] in lab:
                    if all(w in lab for w in want.split()): rowmap[k] = ri
        drec, has_prod = {}, False
        for ci, h in enumerate(heads):
            if h not in LIMITS: continue
            pt = _f(rws[rowmap['prod']][ci]) if 'prod' in rowmap else None
            if not pt or pt <= 0.5: continue
            has_prod = True
            drec[h] = {k: (_f(rws[ri][ci]) if ci < len(rws[ri]) else None)
                       for k, ri in rowmap.items()}
        D['daily'][day] = drec
        if not has_prod: D['zero_days'].append(day)

        # stoppages: ANCHORED — June-style table under a 'stoppage' header,
        # or legacy 'Remarks' free-text lines. Never blind numeric scanning.
        anchored = False
        for ri, r in enumerate(rws):
            if r[0] and 'stoppage' in str(r[0]).lower() and 'duration' in str(r[0]).lower():
                anchored = True
                j = ri + 1
                while j < len(rws) and isinstance(rws[j][0], (int, float)):
                    D['stoppages'].append({'day': day, 'hours': round(float(rws[j][0]), 2),
                                           'dept': str(rws[j][1]).strip() if rws[j][1] else '',
                                           'reason': str(rws[j][2]).strip() if rws[j][2] else ''})
                    j += 1
                break
        if not anchored:
            for ri, r in enumerate(rws):
                if r[0] and str(r[0]).strip().lower() == 'remarks':
                    for j in range(ri + 1, min(ri + 8, len(rws))):
                        txt = next((str(c).strip() for c in rws[j][:4] if c and str(c).strip()), '')
                        if not txt: continue
                        mh = re.search(r'([\d\.]+)\s*(?:h|hr|hour|hours)\b', txt.lower())
                        D['stoppages'].append({'day': day,
                                               'hours': round(float(mh.group(1)), 2) if mh else 0.0,
                                               'dept': '', 'reason': txt[:160]})
                    break

    # ---- fallbacks for missing summary fields (older template) ----
    for p, v in D['products'].items():
        if not v.get('hours'):
            v['hours'] = round(sum((r[p].get('hours') or 0)
                                   for r in D['daily'].values() if p in r), 2) or None
    if not D['plant'].get('hours'):
        net = D['pi'].get('net', {})
        s = sum(v for v in net.values() if v)
        D['plant']['hours'] = round(s, 1) if s else round(sum(
            (v.get('hours') or 0) for v in D['products'].values()), 1)
    if not D['plant'].get('avg_tph') and D['plant'].get('prod') and D['plant'].get('hours'):
        D['plant']['avg_tph'] = round(D['plant']['prod'] / D['plant']['hours'], 2)
    wb.close()
    return D

# --------------------------------------------------------------- ANALYSIS ---
def categorize(s):
    r = s['reason'].lower()
    for cat, test in STOP_CATS:
        if test(r, s['dept']): return cat
    return 'Other production'

def analyze(D, year, month, prev=None, prev2=None, elec_cost=None):
    """prev / prev2: metrics dicts of previous months (may be None)."""
    A = {}
    ndays = calendar.monthrange(year, month)[1]
    A['fridays'] = [d for d in range(1, ndays + 1) if datetime(year, month, d).weekday() == 4]
    A['ndays'] = ndays
    cost = elec_cost if elec_cost else (D['plant'].get('cost') or 0)
    A['cost'] = cost
    A['jd_per_t'] = cost / D['plant']['prod'] if D['plant'].get('prod') else None
    A['tariff'] = cost / D['plant']['kwh'] if D['plant'].get('kwh') else None

    # stoppage categories
    cats, ev = defaultdict(float), defaultdict(int)
    for s in D['stoppages']:
        c = categorize(s); cats[c] += s['hours']; ev[c] += 1
    A['stop_cats'] = sorted(cats.items(), key=lambda x: -x[1])
    A['stop_events'] = dict(ev)
    A['stop_total'] = sum(cats.values())
    A['planned_h'] = sum(v for v in D['pi'].get('planned', {}).values() if v)
    A['incident_h'] = sum(v for v in D['pi'].get('incident', {}).values() if v)
    A['silofull_h'] = cats.get('Silos full / dispatch', 0)
    A['silofull_events'] = [(s['day'], s['hours']) for s in D['stoppages']
                            if categorize(s) == 'Silos full / dispatch']
    avg_tph = D['plant'].get('avg_tph') or 20.0
    A['silofull_loss_t'] = A['silofull_h'] * avg_tph
    A['fan_events'] = [(s['day'], s['hours']) for s in D['stoppages']
                       if categorize(s) == 'Process fan (electrical)']

    # per-product alerts + recipe deviations
    A['alerts'], A['recipe_dev'], A['recipe_norm'] = {}, {}, {}
    for p, (tmin, bmin, bmax, _c) in LIMITS.items():
        rows = {d: r[p] for d, r in D['daily'].items() if p in r}
        if not rows and p in D['products']:
            A['alerts'][p] = {'bl_low': [], 'bl_high': [], 'tph_low': []}
            continue
        A['alerts'][p] = {
         'bl_low':  sorted((d, round(v['blaine'])) for d, v in rows.items() if v.get('blaine') and v['blaine'] < bmin),
         'bl_high': sorted((d, round(v['blaine'])) for d, v in rows.items() if v.get('blaine') and v['blaine'] > bmax),
         'tph_low': sorted((d, round(v['tph'], 2)) for d, v in rows.items() if v.get('tph') and v['tph'] < tmin)}
        stable = [v['clinker'] * 100 for d, v in rows.items()
                  if v.get('clinker') and (v.get('hours') or 0) >= 8]
        pool = stable or [v['clinker'] * 100 for v in rows.values() if v.get('clinker')]
        if pool:
            norm = float(np.median(pool)); A['recipe_norm'][p] = norm
            excess, devs = 0.0, []
            for d, v in sorted(rows.items()):
                if not v.get('clinker'): continue
                cl = v['clinker'] * 100; dv = cl - norm
                if dv > 1.0:
                    t = (v['prod'] or 0) * dv / 100; excess += t
                    devs.append((d, round(cl, 1), round(t, 1)))
                elif dv < -5.0:
                    devs.append((d, round(cl, 1), round((v['prod'] or 0) * dv / 100, 1)))
            A['recipe_dev'][p] = {'excess_t': round(excess, 1), 'days': devs}

    # stock movement + pools
    A['stock'] = {}
    for mat, close in D['stock_close'].items():
        key = mat.strip()
        opn = next((v for k, v in D['stock_open'].items() if k.strip().startswith(key[:10])), None)
        imp = next((v for k, v in D['imports'].items() if k.strip().startswith(key[:10])), 0) or 0
        used = (opn + imp - close) if opn is not None else None
        A['stock'][key] = {'open': opn, 'imports': imp, 'close': close, 'used': used}
    def pool(mats):
        st = sum(A['stock'][m]['close'] for m in mats if m in A['stock'])
        dr = sum(A['stock'][m]['used'] or 0 for m in mats if m in A['stock'])
        return st, dr, (st / dr if dr > 0 else None)
    A['grey_pool'] = pool(GREY_POOL)
    A['white_pool'] = pool(WHITE_POOL)

    # silo reconciliation (M50 bulk rule)
    A['silo_recon'] = None
    if D['silos'].get('start') and D['silos'].get('end'):
        st, en = D['silos']['start'], D['silos']['end']
        diffs = [round((e or 0) - (s or 0), 1) for s, e in zip(st, en)]
        A['silo_recon'] = {'start': st, 'end': en, 'diff': diffs}
        pk = D['silos'].get('packing_raw') or []
        # packing_raw order mirrors summary: M50, -, SW, PW, Eco, CEM, M10
        names = ['M50', None, 'Super white', 'Power white', 'Eco white', 'CEM I 52.5R', 'M10']
        A['packing'] = {n: pk[i] for i, n in enumerate(names) if n and i < len(pk) and pk[i] is not None}
        m50_dsilo = (diffs[0] or 0) + (diffs[1] or 0)
        A['m50_bulk'] = (D['products'].get('M50', {}).get('prod') or 0) - m50_dsilo
        gap = 0.0
        for n, silo_i in [('Super white', 2), ('Power white', 3), ('Eco white', 4)]:
            if n in A['packing'] and n in D['products']:
                gap += D['products'][n]['prod'] - A['packing'][n] - diffs[silo_i]
        for n in ['CEM I 52.5R', 'M10']:
            if n in A['packing'] and n in D['products']:
                gap += D['products'][n]['prod'] - A['packing'][n]
        A['packed_gap'] = round(gap, 1)

    # comparisons
    A['prev'], A['prev2'] = prev, prev2
    return A

def metrics_json(D, A, year, month):
    """Compact per-month metrics stored in DB for future comparisons + /alerts."""
    return {
     'year': year, 'month': month, 'plant': D['plant'], 'cost': A['cost'],
     'jd_per_t': A['jd_per_t'], 'tariff': A['tariff'],
     'products': {p: {k: v.get(k) for k in ('prod', 'tph', 'spc_plant', 'spc_mill', 'blaine', 'wi', 'ck', 'clinker', 'hours')}
                  for p, v in D['products'].items()},
     'planned_h': A['planned_h'], 'incident_h': A['incident_h'],
     'silofull_h': A['silofull_h'], 'stop_cats': A['stop_cats'],
     'alerts': A['alerts'], 'recipe_dev': A['recipe_dev'], 'recipe_norm': A['recipe_norm'],
     'grey_pool': A['grey_pool'], 'white_pool': A['white_pool'],
     'stock': A['stock'], 'zero_days': D['zero_days'], 'missing_days': D['missing_days'],
     'packed_gap': A.get('packed_gap'), 'm50_bulk': A.get('m50_bulk'),
    }

# ----------------------------------------------------------------- CHARTS ---
def make_charts(D, A, ch, year, month):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    plt.rcParams.update({'font.size': 8, 'axes.grid': True, 'grid.alpha': 0.3, 'figure.dpi': 145})
    mn = calendar.month_name[month]
    ndays = A['ndays']
    daily_tot = {d: sum((r[p]['prod'] or 0) for p in r) for d, r in D['daily'].items()}
    days = list(range(1, ndays + 1)); vals = [daily_tot.get(d, 0) for d in days]
    fig, ax = plt.subplots(figsize=(9, 2.9))
    ax.bar(days, vals, color=['#c0392b' if v == 0 else '#2e6db4' for v in vals])
    for f in A['fridays']: ax.axvspan(f - .45, f + .45, color='red', alpha=.06)
    ax.set_xticks(days); ax.set_xlabel(f'Day of {mn}'); ax.set_ylabel('Production (t)')
    ax.set_title(f'Daily Production — {mn} {year}  (red = zero; shaded = Fridays)', fontsize=9)
    plt.tight_layout(); plt.savefig(f'{ch}/daily.png'); plt.close()

    # product comparison vs prev months
    prods = [p for p in LIMITS if p in D['products']]
    series = []
    for tag, mset, col in [('-2', A['prev2'], '#7f8c8d'), ('-1', A['prev'], '#2e6db4'), ('now', None, '#d35400')]:
        if tag == 'now':
            series.append((f'{mn} {year}', [D['products'][p]['prod'] for p in prods], col))
        elif mset:
            lbl = f"{calendar.month_name[mset['month']]} {mset['year']}"
            series.append((lbl, [mset['products'].get(p, {}).get('prod') or 0 for p in prods], col))
    x = np.arange(len(prods)); w = 0.8 / max(len(series), 1)
    fig, ax = plt.subplots(figsize=(9, 3.1))
    for i, (lbl, ys, col) in enumerate(series):
        ax.bar(x + (i - (len(series) - 1) / 2) * w, ys, w, label=lbl, color=col)
    for i, v in enumerate(series[-1][1]):
        ax.text(i + (len(series) - 1) / 2 * w, v + 40, f'{v:,.0f}', ha='center', fontsize=7, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels([p.replace(' ', '\n') for p in prods]); ax.legend(); ax.set_ylabel('Tons')
    ax.set_title('Production by Product — Month Comparison', fontsize=9)
    plt.tight_layout(); plt.savefig(f'{ch}/prod_cmp.png'); plt.close()

    # SPC comparison
    fig, ax = plt.subplots(figsize=(9, 2.9))
    for i, (lbl, _ys, col) in enumerate(series):
        if lbl.startswith(mn) and i == len(series) - 1:
            ys = [D['products'][p]['spc_plant'] for p in prods]
        else:
            mset = A['prev2'] if i == 0 and len(series) == 3 else A['prev']
            ys = [mset['products'].get(p, {}).get('spc_plant') or 0 for p in prods]
        ax.bar(x + (i - (len(series) - 1) / 2) * w, ys, w, label=lbl, color=col)
    ax.set_xticks(x); ax.set_xticklabels([p.replace(' ', '\n') for p in prods]); ax.legend(loc='upper left')
    ax.set_ylabel('kWh/t'); ax.set_title('SPC Plant by Product — Month Comparison', fontsize=9)
    plt.tight_layout(); plt.savefig(f'{ch}/spc_cmp.png'); plt.close()

    # stoppage pareto
    cats = A['stop_cats']
    fig, ax = plt.subplots(figsize=(9, 3.1))
    names = [c for c, _ in cats]; hrs = [h for _, h in cats]
    cols = ['#c0392b' if 'Silos' in n else '#95a5a6' if 'holiday' in n else '#e67e22' if 'fan' in n else '#7f8c8d' for n in names]
    bars = ax.bar(names, hrs, color=cols)
    for b, h in zip(bars, hrs): ax.text(b.get_x() + b.get_width() / 2, h + .5, f'{h:.1f}h', ha='center', fontsize=7.5, fontweight='bold')
    if sum(hrs):
        ax2 = ax.twinx(); ax2.plot(names, np.cumsum(hrs) / sum(hrs) * 100, 'k--o', ms=3, lw=1)
        ax2.set_ylabel('Cum %'); ax2.set_ylim(0, 105); ax2.grid(False)
    ax.set_ylabel('Hours'); ax.set_title(f'Stoppage Root-Cause Pareto — {mn} {year} ({sum(hrs):.1f} h logged)', fontsize=9)
    plt.setp(ax.get_xticklabels(), rotation=13, ha='right', fontsize=7)
    plt.tight_layout(); plt.savefig(f'{ch}/pareto.png'); plt.close()

    # silo-full timeline
    if A['silofull_events']:
        fig, ax = plt.subplots(figsize=(9, 2.2))
        ax.bar([d for d, _ in A['silofull_events']], [h for _, h in A['silofull_events']], color='#c0392b', width=.6)
        for d, h in A['silofull_events']: ax.text(d, h + .2, f'{h}', ha='center', fontsize=7)
        ax.set_xticks(range(1, ndays + 1)); ax.set_ylabel('Hours')
        ax.set_title(f'"Silos Full" Stoppages — {A["silofull_h"]:.1f} h (~{A["silofull_loss_t"]:.0f} t lost potential)', fontsize=9)
        plt.tight_layout(); plt.savefig(f'{ch}/silofull.png'); plt.close()

    # per-product charts
    for p, (tmin, bmin, bmax, color) in LIMITS.items():
        rows = {d: r[p] for d, r in D['daily'].items() if p in r}
        if not rows: continue
        ds = sorted(rows); xs = np.arange(len(ds))
        key = p.replace(' ', '_').replace('.', '')
        tph = [rows[d]['tph'] or 0 for d in ds]
        fig, ax = plt.subplots(figsize=(8.6, 1.9))
        ax.plot(xs, tph, '-o', ms=3.5, color=color)
        ax.axhline(tmin, ls='--', color='red', lw=.9, label=f'Min {tmin}')
        ax.axhline(np.mean(tph), ls=':', color='gray', lw=.9, label=f'Avg {np.mean(tph):.2f}')
        ax.set_xticks(xs); ax.set_xticklabels(ds); ax.set_ylabel('t/h'); ax.legend(fontsize=6.5)
        plt.tight_layout(); plt.savefig(f'{ch}/{key}_tph.png'); plt.close()
        bl = [rows[d]['blaine'] or 0 for d in ds]
        fig, ax = plt.subplots(figsize=(8.6, 1.9))
        ax.bar(xs, bl, .55, color=['#c0392b' if (v < bmin or v > bmax) else color for v in bl])
        ax.axhline(bmin, ls='--', color='red', lw=.9, label=f'Min {bmin}')
        ax.axhline(bmax, ls='--', color='green', lw=.9, label=f'Max {bmax}')
        ax.set_ylim(min(bl + [bmin]) - 250, max(bl + [bmax]) + 150)
        ax.set_xticks(xs); ax.set_xticklabels(ds); ax.set_ylabel('cm2/g'); ax.legend(fontsize=6.5, loc='lower right')
        plt.tight_layout(); plt.savefig(f'{ch}/{key}_blaine.png'); plt.close()

    # recipe deviation chart (products having deviation days)
    devp = [p for p, dv in A['recipe_dev'].items() if dv['days']]
    if devp:
        fig, axes = plt.subplots(len(devp), 1, figsize=(9, 2.3 * len(devp)), squeeze=False)
        for ax, p in zip(axes[:, 0], devp):
            norm = A['recipe_norm'][p]
            rows = {d: r[p] for d, r in D['daily'].items() if p in r}
            ds = sorted(rows); cl = [(rows[d]['clinker'] or 0) * 100 for d in ds]
            xs = np.arange(len(ds))
            ax.bar(xs, cl, .6, color=['#c0392b' if c > norm + 1 else '#e67e22' if c < norm - 5 else LIMITS[p][3] for c in cl])
            ax.axhline(norm, ls='--', color='black', lw=1, label=f'Norm {norm:.1f}%')
            ax.set_xticks(xs); ax.set_xticklabels(ds)
            ax.set_ylim(min(cl) - 6, max(cl) + 6); ax.set_ylabel('Clinker %'); ax.legend(fontsize=7, loc='lower right')
            ax.set_title(f'{p} — daily clinker ratio (red = excess clinker)', fontsize=8.5)
        plt.tight_layout(); plt.savefig(f'{ch}/recipe.png'); plt.close()

    # stock chart
    mats = sorted(A['stock'].items(), key=lambda x: -(x[1]['close'] or 0))
    fig, ax = plt.subplots(figsize=(8.6, 2.9))
    names = [m for m, _ in mats][::-1]; v = [(x['close'] or 0) for _, x in mats][::-1]
    ax.barh(names, v, color=['#c0392b' if x < 500 else '#e67e22' if x < 1100 else '#2e6db4' for x in v])
    for i, x in enumerate(v): ax.text(x + max(v) * .01, i, f'{x:,.0f}', va='center', fontsize=7)
    ax.set_xlabel('Tons'); ax.set_title(f'Final Raw Material Stock — {calendar.month_name[month]} {year}', fontsize=9)
    plt.tight_layout(); plt.savefig(f'{ch}/stock.png'); plt.close()


# --------------------------------------------- RULE-BASED FALLBACK CONTENT ---
def _rule_insights(D, A, P, PL, prev, wc, gc):
    """Deterministic insights used when the AI layer is unavailable."""
    ins = []
    if A['silofull_h'] > 10:
        ins.append(('The bottleneck is dispatch, not the mill.',
                    f"{A['silofull_h']:.1f} h of silos-full stops (~{A['silofull_loss_t']:.0f} t) vs "
                    f"{A['incident_h']:.1f} h of incidents. Improvement effort returns more on the dispatch chain."))
    drift = [p for p in P if A['alerts'].get(p, {}).get('bl_low')
             and len(A['alerts'][p]['bl_low']) >= max(2, len([d for d, r in D['daily'].items() if p in r]) // 3)]
    if drift:
        ins.append(('Fineness margin is being consumed.',
                    'Repeated Blaine-below-min days on: ' + ', '.join(drift) +
                    '. If SPC gains coincide with falling Blaine, part of the saving is coarser grinding.'))
    slow = []
    for p in P:
        rows_ = {d: r[p] for d, r in D['daily'].items() if p in r}
        for d, _v in A['alerts'].get(p, {}).get('tph_low', []):
            if (rows_.get(d, {}).get('hours') or 24) < 12: slow.append((p, d))
    if slow:
        ins.append(('Low-productivity days are restart days.',
                    'Sub-minimum t/h days (' + ', '.join(f'{p} D{d}' for p, d in slow[:6]) +
                    ') are short runs after stoppages — a restart-frequency problem.'))
    if wc is not None and wc < WHITE_FLOOR:
        ins.append(('White clinker coverage is the binding constraint.',
                    f"White pool at {wc:.1f} months vs {WHITE_FLOOR:.0f}-month floor — procurement lead time now "
                    f"gates the white portfolio."))
    tot_ex = sum(rd['excess_t'] for rd in A['recipe_dev'].values())
    if tot_ex > 5:
        ins.append(('Stops cost clinker, not just hours.',
                    f"~{tot_ex:.0f} t of excess clinker traces to unstabilized post-stoppage starts."))
    if prev:
        ins.append(('Effective capacity estimate.',
                    f"Removing dispatch-driven stops lifts the month to ~{PL['prod'] + A['silofull_loss_t']:,.0f} t "
                    f"(actual {PL['prod']:,.0f} t)."))
    if not ins:
        ins.append(('Stable month.', 'No cross-linked anomalies detected by the rule engine.'))
    return ins


def _rule_recs(D, A, P, PL, prev, wc, gc, ws_, wd, gs, gd, bl_drift, tot_excess2, pct, prev_name):
    recs = [['Priority', 'Recommendation']]
    if A['silofull_h'] > 10:
        recs.append(['URGENT', f"Debottleneck packing & dispatch — {A['silofull_h']:.1f} h of silos-full stops "
                               f"(~{A['silofull_loss_t']:.0f} t)."])
    if wc is not None and wc < WHITE_FLOOR:
        recs.append(['URGENT', f"White clinker pool at {wc:.1f} months (< {WHITE_FLOOR:.0f}) — order "
                               f"~{WHITE_FLOOR*wd-ws_:,.0f} t now."])
    if gc is not None and gc < GREY_FLOOR + 0.5:
        recs.append(['HIGH', f"Grey clinker pool at {gc:.1f} months (floor {GREY_FLOOR:.0f}) — schedule replenishment."])
    if len(A['fan_events']) >= 2:
        recs.append(['HIGH', f"Process-fan RCA — {len(A['fan_events'])} recurrences "
                             f"({sum(h for _, h in A['fan_events']):.1f} h)."])
    for p in bl_drift:
        recs.append(['HIGH', f"Correct {p} fineness drift — repeated Blaine-below-min days."])
    if tot_excess2 > 5:
        recs.append(['MEDIUM', f"Recipe governance — ~{tot_excess2:.0f} t excess clinker on restart transients."])
    if A.get('packed_gap') and abs(A['packed_gap']) > 20:
        recs.append(['MEDIUM', f"Investigate {abs(A['packed_gap']):,.0f} t packed-products reconciliation gap."])
    if D['missing_days']:
        recs.append(['MEDIUM', f"Recover missing daily sheets: {D['missing_days']}."])
    if prev and PL['prod'] > (prev['plant'].get('prod') or 0):
        recs.append(['POSITIVE', f"Production {pct(PL['prod'], prev['plant'].get('prod'))} vs {prev_name} at "
                                 f"utilization {(PL.get('utilization') or 0)*100:.1f}%."])
    if len(recs) == 1:
        recs.append(['—', 'No actions triggered by the rule engine this month.'])
    return recs


# -------------------------------------------------------------------- PDF ---
def _styles():
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors
    BLUE = colors.HexColor(BLUE_HEX); DBLUE = colors.HexColor(DBLUE_HEX); RED = colors.HexColor(RED_HEX)
    base = getSampleStyleSheet()['Normal']
    def st(name, **kw): return ParagraphStyle(name, parent=base, **kw)
    return {
     'h1': st('h1', fontSize=15, textColor=BLUE, fontName='Helvetica-Bold', spaceAfter=4),
     'h2': st('h2', fontSize=11, fontName='Helvetica-Bold', spaceAfter=3, spaceBefore=6),
     'body': st('body', fontSize=8.4, leading=11.5),
     'small': st('small', fontSize=7.4, leading=9.6),
     'kpi_v': st('kpi_v', fontSize=15, fontName='Helvetica-Bold', alignment=TA_CENTER),
     'kpi_u': st('kpi_u', fontSize=7, alignment=TA_CENTER, textColor=colors.HexColor('#666666')),
     'kpi_l': st('kpi_l', fontSize=7.5, alignment=TA_CENTER, textColor=colors.HexColor('#444444')),
     'alert_red': st('ar', fontSize=8.2, leading=11, textColor=RED, fontName='Helvetica-Bold'),
     'alert_orange': st('ao', fontSize=8.2, leading=11, textColor=colors.HexColor('#9a5b00'), fontName='Helvetica-Bold'),
     'alert_green': st('ag', fontSize=8.2, leading=11, textColor=colors.HexColor('#1d6f42'), fontName='Helvetica-Bold'),
     'ins': st('ins', fontSize=8.8, textColor=DBLUE, leading=11, fontName='Helvetica-Bold'),
    }

def build_pdf(D, A, ch, out_path, year, month, ai=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (BaseDocTemplate, PageTemplate, Frame, Paragraph,
                                    Spacer, Table, TableStyle, Image, PageBreak)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    BLUE = colors.HexColor(BLUE_HEX); DBLUE = colors.HexColor(DBLUE_HEX)
    LGREY = colors.HexColor(LGREY_HEX); RED = colors.HexColor(RED_HEX)
    ORANGE = colors.HexColor(ORANGE_HEX); GREEN = colors.HexColor(GREEN_HEX)
    S = _styles(); W, H = A4
    mn = calendar.month_name[month]
    title = f'Monthly Production Report — {mn} {year}'

    def hf(canvas, doc):
        canvas.saveState(); canvas.setFillColor(DBLUE)
        canvas.rect(0, H - 14 * mm, W, 14 * mm, fill=1, stroke=0)
        canvas.setFillColor(colors.white); canvas.setFont('Helvetica-Bold', 9.5)
        canvas.drawString(14 * mm, H - 9 * mm, f'{title}  |  Auto-generated by SigmaCement_bot')
        canvas.drawRightString(W - 14 * mm, H - 9 * mm, f'Page {doc.page}')
        canvas.setFillColor(colors.HexColor('#eef1f4')); canvas.rect(0, 0, W, 11 * mm, fill=1, stroke=0)
        canvas.setFillColor(colors.HexColor('#777777')); canvas.setFont('Helvetica', 7.5)
        canvas.drawString(14 * mm, 4.5 * mm, 'Confidential — Internal Use Only')
        canvas.drawRightString(W - 14 * mm, 4.5 * mm, 'Production Monitoring System © 2026')
        canvas.restoreState()

    doc = BaseDocTemplate(out_path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                          topMargin=20 * mm, bottomMargin=15 * mm)
    doc.addPageTemplates([PageTemplate(id='p', frames=[Frame(14 * mm, 15 * mm, W - 28 * mm, H - 38 * mm)], onPage=hf)])
    E = []
    CW = W - 28 * mm

    def kpi(items):
        cells = [Paragraph(f'<font color="{c}">{v}</font>', S['kpi_v']) for v, u, l, c in items]
        r2 = [Paragraph(u, S['kpi_u']) for v, u, l, c in items]
        r3 = [Paragraph(l, S['kpi_l']) for v, u, l, c in items]
        t = Table([cells, r2, r3], colWidths=[CW / len(items)] * len(items))
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), LGREY),
                               ('BOX', (0, 0), (-1, -1), .8, colors.HexColor('#9fb6cc')),
                               ('INNERGRID', (0, 0), (-1, 0), .4, colors.HexColor('#c6d3df')),
                               ('TOPPADDING', (0, 0), (-1, 0), 6), ('BOTTOMPADDING', (0, 2), (-1, 2), 5)]))
        E.append(t); E.append(Spacer(1, 4))

    def alert(text, kind='red'):
        bg = {'red': '#fdecea', 'orange': '#fdf3e3', 'green': '#e9f7ef'}[kind]
        bd = {'red': RED, 'orange': ORANGE, 'green': GREEN}[kind]
        t = Table([[Paragraph(text, S['alert_' + kind])]], colWidths=[CW])
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(bg)),
                               ('BOX', (0, 0), (-1, -1), .8, bd), ('LEFTPADDING', (0, 0), (-1, -1), 7),
                               ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4)]))
        E.append(t); E.append(Spacer(1, 2.5))

    def tbl(data, colw=None, fs=6.8, align='CENTER', hdr=1):
        cst = ParagraphStyle('c', fontSize=fs, alignment=TA_CENTER if align == 'CENTER' else TA_LEFT, leading=fs + 1.6)
        td = [[Paragraph(f'<b><font color="white">{c}</font></b>' if ri < hdr else str(c), cst) for c in row]
              for ri, row in enumerate(data)]
        t = Table(td, colWidths=colw, repeatRows=hdr)
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, hdr - 1), BLUE),
                               ('GRID', (0, 0), (-1, -1), .4, colors.HexColor('#b9c6d2')),
                               ('ROWBACKGROUNDS', (0, hdr), (-1, -1), [colors.white, LGREY]),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                               ('TOPPADDING', (0, 0), (-1, -1), 2.2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.2)]))
        E.append(t); E.append(Spacer(1, 4))

    def img(path, w=None):
        if not os.path.exists(path): return
        from PIL import Image as PIL
        iw, ih = PIL.open(path).size
        w = w or CW
        E.append(Image(path, width=w, height=w * ih / iw)); E.append(Spacer(1, 3))

    def sec(num, t):
        E.append(Spacer(1, 4)); E.append(Paragraph(f'{num}. {t}', S['h1']))
        bar = Table([['']], colWidths=[CW], rowHeights=[1.2])
        bar.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), BLUE)]))
        E.append(bar); E.append(Spacer(1, 5))

    def pct(a, b):
        return f'{(a / b - 1) * 100:+.1f}%' if (a is not None and b) else '—'
    def fnum(v, f='{:,.1f}'):
        return f.format(v) if v is not None else '—'

    P, PL = D['products'], D['plant']
    prev, prev2 = A['prev'], A['prev2']
    pv = lambda p, k: (prev or {}).get('products', {}).get(p, {}).get(k) if prev else None
    pv2 = lambda p, k: (prev2 or {}).get('products', {}).get(p, {}).get(k) if prev2 else None
    prev_name = f"{calendar.month_name[prev['month']]}" if prev else 'prev month'

    # ===== 1 EXEC =====
    def g(v, f='{:,.1f}'):
        return f.format(v) if v is not None else '—'
    def gp(v):
        return f'{v*100:.1f}%' if v is not None else '—'
    sec('1', 'Executive Summary')
    if ai and ai.get('headline'):
        hb = Table([[Paragraph(f"<b>{ai['headline']}</b>", S['ins'])]], colWidths=[CW])
        hb.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#eaf1fa')),
                                ('BOX', (0, 0), (-1, -1), .9, BLUE), ('LEFTPADDING', (0, 0), (-1, -1), 8),
                                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6)]))
        E.append(hb); E.append(Spacer(1, 5))
    kpi([(g(PL.get('prod')), 't', 'Total Production', '#164e87'),
         (g(PL.get('hours'), '{:.1f}'), 'h/month', 'Running Hours', '#1d6f42'),
         (g(PL.get('avg_tph'), '{:.2f}'), 't/h', 'Avg Mill Productivity', '#b9770e'),
         (g(PL.get('kwh'), '{:,.0f}'), 'kWh', 'Power Consumption', '#8e44ad')])
    kpi([(g(PL.get('spc'), '{:.2f}'), 'kWh/t', f"Plant SPC ({prev_name}: {fnum((prev or {}).get('plant',{}).get('spc'),'{:.2f}') if prev else '—'})", '#1d6f42'),
         (f"{A['cost']:,.0f}", 'JD/month', f"Electricity Cost ({pct(A['cost'], (prev or {}).get('cost')) if prev else '—'})", '#117864'),
         (gp(PL.get('availability')), 'overall', 'Availability', '#1a5fa8'),
         (gp(PL.get('utilization')), '', 'Utilization', '#1a5fa8')])
    zp = ', '.join(str(d) for d in D['zero_days']) or 'none'
    all_fri = all(d in A['fridays'] for d in D['zero_days'])
    if prev and PL['prod'] > (prev['plant'].get('prod') or 0):
        alert(f"Production {pct(PL['prod'], prev['plant'].get('prod'))} vs {prev_name}"
              + (f" and {pct(PL['prod'], prev2['plant'].get('prod'))} vs {calendar.month_name[prev2['month']]}" if prev2 else '')
              + f". Zero-production days: {zp}" + (' — all Fridays (weekly planned stops).' if all_fri and D['zero_days'] else '.'), 'green')
    else:
        alert(f"Zero-production days: {zp}" + (' — all Fridays.' if all_fri and D['zero_days'] else '.'), 'orange')
    if D['missing_days']:
        alert(f"MISSING daily sheets for day(s): {D['missing_days']} — verify actual status before treating as zero (house rule).", 'red')
    if prev:
        alert(f"Incident downtime {A['incident_h']:.1f} h vs {(prev.get('incident_h') or 0):.1f} h in {prev_name}. "
              f"Top actionable loss: silos-full stops {A['silofull_h']:.1f} h (~{A['silofull_loss_t']:.0f} t potential). See Section 2.", 'orange')
    img(f'{ch}/daily.png'); img(f'{ch}/prod_cmp.png')
    E.append(Paragraph('Product Performance Summary', S['h2']))
    rowsx = [['Product', 'Prod (t)', f'{prev_name} (t)', 'Change', 't/h', 'SPC', f'{prev_name} SPC', 'Blaine']]
    for p in LIMITS:
        if p not in P: continue
        rowsx.append([p, f"{P[p]['prod']:,.1f}", fnum(pv(p, 'prod'), '{:,.1f}'), pct(P[p]['prod'], pv(p, 'prod')),
                      f"{P[p]['tph']:.2f}", f"{P[p]['spc_plant']:.2f}", fnum(pv(p, 'spc_plant'), '{:.2f}'),
                      f"{P[p]['blaine']:.0f}"])
    tbl(rowsx)
    E.append(Paragraph('Production totals from the Summary sheet (authoritative). "Power white-R" mapped to CEM I 52.5R and '
                       '"Super white Special" to M10 per naming rules; all clinker variants consolidated as Total Clinker.', S['small']))
    E.append(PageBreak())

    # ===== 2 STOPPAGES =====
    sec('2', 'Stoppages & Downtime Study')
    kpi([(f"{A['planned_h']+A['incident_h']:.1f}", 'hours', 'Total downtime', '#c0392b'),
         (f"{A['planned_h']:.1f}", 'hours', 'Planned', '#1a5fa8'),
         (f"{A['incident_h']:.1f}", 'hours', f"Incidents ({prev_name}: {fnum((prev or {}).get('incident_h'),'{:.1f}') if prev else '—'})", '#c0392b'),
         (f"{A['silofull_h']:.1f}", 'hours', 'Silos-full stops', '#c0392b')])
    img(f'{ch}/pareto.png')
    ct = [['Root cause', 'Hours', 'Share', 'Events']]
    tot = A['stop_total'] or 1
    for c, h in A['stop_cats']:
        ct.append([c, f'{h:.1f}', f'{h/tot*100:.1f}%', A['stop_events'].get(c, 0)])
    tbl(ct, colw=[90 * mm, 25 * mm, 25 * mm, 42 * mm], align='LEFT')
    E.append(Paragraph('Full Stoppage Log', S['h2']))
    lg = [['Day', 'h', 'Dept', 'Reason (as logged)']]
    for s in sorted(D['stoppages'], key=lambda x: x['day']):
        lg.append([s['day'], f"{s['hours']:.2f}", s['dept'], s['reason'][:110]])
    tbl(lg, colw=[9 * mm, 11 * mm, 20 * mm, 142 * mm], fs=6.3, align='LEFT')
    if A['silofull_h'] > 5:
        E.append(Paragraph('Deep Dive — "Silos Full" (dispatch constraint)', S['h2']))
        img(f'{ch}/silofull.png')
        alert(f"The mill stopped {A['silofull_h']:.1f} h because product silos were full — ~{A['silofull_loss_t']:.0f} t of lost "
              f"potential at {g(PL.get('avg_tph'), '{:.2f}')} t/h. The constraint is packing/dispatch capacity, not the mill: "
              f"align dispatch with high-production days and pre-draw silos before weekends.", 'red')
    if len(A['fan_events']) >= 2:
        alert(f"Recurring process-fan fault: {len(A['fan_events'])} events on days {[d for d,_ in A['fan_events']]} "
              f"({sum(h for _,h in A['fan_events']):.1f} h). If events recur after an intervention, escalate to a formal RCA "
              f"on the fan drive / PID loop.", 'orange')
    E.append(PageBreak())

    # ===== 3 POWER =====
    sec('3', 'Power Consumption')
    img(f'{ch}/spc_cmp.png')
    pt = [['Product', 'SPC Plant', f'{prev_name} SPC', 'Change', 'SPC Mill']]
    for p in LIMITS:
        if p not in P: continue
        pt.append([p, f"{P[p]['spc_plant']:.2f}", fnum(pv(p, 'spc_plant'), '{:.2f}'),
                   pct(P[p]['spc_plant'], pv(p, 'spc_plant')), f"{P[p]['spc_mill']:.2f}"])
    tbl(pt)
    msg = (f"Electricity cost {A['cost']:,.0f} JD"
           + (f" ({pct(A['cost'], prev.get('cost'))} vs {prev_name})" if prev and prev.get('cost') else '')
           + f" | cost per ton {A['jd_per_t']:.2f} JD/t"
           + (f" ({prev_name}: {prev.get('jd_per_t'):.2f})" if prev and prev.get('jd_per_t') else '')
           + f" | effective tariff {A['tariff']:.4f} JD/kWh"
           + (f" ({prev_name}: {prev.get('tariff'):.4f})" if prev and prev.get('tariff') else '') + '.')
    good = (not prev) or (not prev.get('jd_per_t')) or (A['jd_per_t'] or 9e9) <= prev['jd_per_t'] * 1.02
    alert(msg, 'green' if good else 'orange')
    E.append(PageBreak())

    # ===== 4 PRODUCTS =====
    for p in LIMITS:
        if p not in P: continue
        tmin, bmin, bmax, color = LIMITS[p]
        sec('4', f'Product: {p}')
        kpi([(f"{P[p]['prod']:,.1f}", 't', 'Total Production', color),
             (f"{P[p]['hours']:.1f}", 'h', 'Running Hours', color),
             (f"{P[p]['tph']:.2f}", 't/h', 'Avg Productivity', color),
             (f"{P[p]['spc_plant']:.2f}", 'kWh/t', 'Avg Plant SPC', color)])
        key = p.replace(' ', '_').replace('.', '')
        rows = {d: r[p] for d, r in D['daily'].items() if p in r}
        img(f'{ch}/{key}_tph.png')
        if rows:
            ds = sorted(rows)
            def mrow(k, fmt, scale=1):
                return [fmt.format((rows[d][k] or 0) * scale) if rows[d].get(k) is not None else '—' for d in ds]
            data = [['Metric'] + [str(d) for d in ds],
                    ['t/h'] + mrow('tph', '{:.2f}'), ['SPC Plant'] + mrow('spc_plant', '{:.1f}'),
                    ['SPC Mill'] + mrow('spc_mill', '{:.1f}'), ['Blaine'] + mrow('blaine', '{:.0f}'),
                    ['R45 %'] + mrow('r45', '{:.2f}'), ['C/K'] + mrow('ck', '{:.3f}'),
                    ['Clinker %'] + mrow('clinker', '{:.1f}', 100)]
            wi = mrow('wi', '{:.2f}')
            if any(x not in ('—', '0.00') for x in wi): data.append(['WI %'] + wi)
            def rend(t):
                n = len(t[0]); lw = 16 * mm
                tbl(t, colw=[lw] + [(CW - lw) / (n - 1)] * (n - 1), fs=6.2)
            if len(ds) > 13:
                half = (len(ds) + 1) // 2
                rend([[r[0]] + r[1:1 + half] for r in data]); rend([[r[0]] + r[1 + half:] for r in data])
            else:
                rend(data)
        img(f'{ch}/{key}_blaine.png')
        E.append(Paragraph(f'Month-on-Month — {p}', S['h2']))
        mom = [['Metric'] + ([f"{calendar.month_name[prev2['month']]}"] if prev2 else []) + ([prev_name] if prev else []) + [mn, 'Chg']]
        for lbl, k, f in [('Production (t)', 'prod', '{:,.1f}'), ('Avg t/h', 'tph', '{:.2f}'),
                          ('SPC Plant', 'spc_plant', '{:.2f}'), ('Blaine', 'blaine', '{:.0f}')]:
            mom.append([lbl] + ([fnum(pv2(p, k), f)] if prev2 else []) + ([fnum(pv(p, k), f)] if prev else [])
                       + [f.format(P[p][k]) if P[p].get(k) is not None else '—', pct(P[p].get(k), pv(p, k))])
        tbl(mom)
        E.append(Paragraph('Observations & Alerts (this product)', S['h2']))
        al = A['alerts'].get(p, {})
        if al.get('bl_low') or al.get('bl_high'):
            parts = []
            if al.get('bl_low'): parts.append(f"below min ({bmin}) on " + ', '.join(f'Day {d} ({v})' for d, v in al['bl_low']))
            if al.get('bl_high'): parts.append(f"above max ({bmax}) on " + ', '.join(f'Day {d} ({v})' for d, v in al['bl_high']))
            avg_flag = ' Monthly average is BELOW minimum.' if P[p]['blaine'] and P[p]['blaine'] < bmin else ''
            alert('ALERT — Blaine ' + '; '.join(parts) + '.' + avg_flag, 'red')
        if al.get('tph_low'):
            short = all((rows.get(d, {}).get('hours') or 24) < 12 for d, _ in al['tph_low']) if rows else False
            alert(f"WATCH — t/h below min ({tmin}) on " + ', '.join(f'Day {d} ({v})' for d, v in al['tph_low'])
                  + ('. All are short/restart runs — start-up drag, not capability loss.' if short else '.'), 'orange')
        rd = A['recipe_dev'].get(p)
        if rd and rd['days']:
            alert(f"Recipe deviations vs steady-state norm {A['recipe_norm'][p]:.1f}%: "
                  + ', '.join(f'Day {d} ({c}%)' for d, c, _t in rd['days'])
                  + (f". Excess clinker consumed: ~{rd['excess_t']:.0f} t." if rd['excess_t'] > 1 else '.'), 'orange')
        if not (al.get('bl_low') or al.get('bl_high') or al.get('tph_low') or (rd and rd['days'])):
            alert('OK — no quality or productivity threshold breaches this month.', 'green')
        if p in BULK_PRODUCTS:
            alert('Note: this product ships 100% in bulk — no packing line applies.', 'green')
        E.append(PageBreak())

    # ===== 5 RECIPE =====
    sec('5', 'Production Recipe Analysis')
    rt = [['Product', 'Clinker %', 'Limestone %', 'Gypsum %', 'Pozzolana %', 'C/K', f'{prev_name} Clinker %']]
    for p in LIMITS:
        if p not in P: continue
        pcl = pv(p, 'clinker')
        rt.append([p, f"{(P[p]['clinker'] or 0)*100:.1f}", f"{(P[p]['limestone'] or 0)*100:.1f}",
                   f"{(P[p]['gypsum'] or 0)*100:.1f}",
                   f"{(P[p]['pozzolana'] or 0)*100:.1f}" if P[p].get('pozzolana') else '—',
                   f"{P[p]['ck']:.3f}" if P[p].get('ck') and P[p]['ck'] < 3 else '(check formula)',
                   f'{pcl*100:.1f}' if pcl else '—'])
    tbl(rt)
    img(f'{ch}/recipe.png')
    tot_excess = sum(rd['excess_t'] for rd in A['recipe_dev'].values())
    if tot_excess > 5:
        alert(f"EXCESS CLINKER: ~{tot_excess:.0f} t consumed above steady-state recipes, concentrated on short post-stoppage "
              f"runs — clinker is the most expensive input; stabilize additive feeders before counting production after restarts.", 'red')
    else:
        alert('Recipes well controlled — no material excess-clinker consumption this month.', 'green')
    E.append(PageBreak())

    # ===== 6 STOCK =====
    sec('6', 'Raw Material Stock & Clinker Pool Policy')
    img(f'{ch}/stock.png')
    stt = [['Material', 'Close (t)', 'Open (t)', 'Imports (t)', 'Consumed (t)']]
    for mat, v in sorted(A['stock'].items(), key=lambda x: -(x[1]['close'] or 0)):
        stt.append([mat, fnum(v['close'], '{:,.0f}'), fnum(v['open'], '{:,.0f}'),
                    fnum(v['imports'], '{:,.0f}'), fnum(v['used'], '{:,.0f}')])
    tbl(stt, colw=[46 * mm, 34 * mm, 34 * mm, 34 * mm, 34 * mm])
    gs, gd, gc = A['grey_pool']; ws_, wd, wc = A['white_pool']
    pool_t = [['Clinker pool', 'Used for', 'Stock (t)', 'Monthly draw (t)', 'Coverage', 'Policy floor'],
              ['Clinker J + M', 'M50 (grey)', f'{gs:,.0f}', f'{gd:,.0f}', f'{gc:.1f} months' if gc else '—', f'{GREY_FLOOR:.0f} months'],
              ['Clinker ALB + SFW + RAK', 'All whites', f'{ws_:,.0f}', f'{wd:,.0f}', f'{wc:.1f} months' if wc else '—', f'{WHITE_FLOOR:.0f} months']]
    tbl(pool_t)
    if wc is not None:
        if wc < WHITE_FLOOR:
            alert(f"BELOW POLICY — white pool at {wc:.1f} months (< {WHITE_FLOOR:.0f}). Order ~{WHITE_FLOOR*wd-ws_:,.0f} t "
                  f"of white clinker now.", 'red')
        else:
            alert(f"White pool at {wc:.1f} months — within policy.", 'green')
    if gc is not None:
        if gc < GREY_FLOOR:
            alert(f"BELOW POLICY — grey pool at {gc:.1f} months (< {GREY_FLOOR:.0f}). Order ~{GREY_FLOOR*gd-gs:,.0f} t of grey clinker.", 'red')
        elif gc < GREY_FLOOR + 0.5:
            alert(f"AT POLICY FLOOR — grey pool at {gc:.1f} months with minimal buffer; schedule replenishment this month.", 'orange')
        else:
            alert(f"Grey pool at {gc:.1f} months — within policy.", 'green')
    dead = [m for m, v in A['stock'].items() if v['used'] is not None and abs(v['used']) < 1 and (v['close'] or 0) > 5 and 'M' != m.split()[-1]]
    if dead:
        alert('No movement this month (possible dead stock): ' + ', '.join(dead) + '.', 'orange')
    E.append(PageBreak())

    # ===== 7 SILOS =====
    if A.get('silo_recon'):
        sec('7', 'Silo Levels & Reconciliation')
        sr = A['silo_recon']
        st_ = [['Silo', 'Start (t)', 'End (t)', 'Diff (t)']]
        for i in range(5):
            st_.append([f'Silo {i+1}', fnum(sr['start'][i], '{:,.1f}'), fnum(sr['end'][i], '{:,.1f}'), fnum(sr['diff'][i], '{:+,.1f}')])
        tbl(st_, colw=[30 * mm, 50 * mm, 50 * mm, 52 * mm])
        pk = [['Product', 'Production (t)', 'Packed / out (t)', 'Unaccounted (t)']]
        for p in LIMITS:
            if p not in P: continue
            if p in BULK_PRODUCTS:
                pk.append([p, f"{P[p]['prod']:,.1f}", f"Bulk — no packing (implied out ~{A['m50_bulk']:,.0f})", '—'])
            elif p in A.get('packing', {}):
                silo_i = {'Super white': 2, 'Power white': 3, 'Eco white': 4}.get(p)
                gap = P[p]['prod'] - A['packing'][p] - (sr['diff'][silo_i] if silo_i is not None else 0)
                pk.append([p, f"{P[p]['prod']:,.1f}", f"{A['packing'][p]:,.0f}", f'{gap:,.1f}'])
        tbl(pk, colw=[36 * mm, 40 * mm, 66 * mm, 40 * mm])
        if A.get('packed_gap') is not None and abs(A['packed_gap']) > 20:
            alert(f"{abs(A['packed_gap']):,.1f} t unaccounted between production, packing and silo movements on PACKED products "
                  f"(M50 bulk excluded). Check silo level calibration and packing records.", 'red')
        E.append(PageBreak())

    # ===== 8 INSIGHTS =====
    bl_drift = [p for p in P if A['alerts'].get(p, {}).get('bl_low')
                and len(A['alerts'][p]['bl_low']) >= max(2, len([d for d, r in D['daily'].items() if p in r]) // 3)]
    tot_excess2 = sum(rd['excess_t'] for rd in A['recipe_dev'].values())
    sec('8', 'Analytical Insights — Cross-Linked Findings')
    if ai and ai.get('insights'):
        E.append(Paragraph('Findings produced by linking the daily stoppage log, power register, stock and silo '
                           'movements, recipe ratios and quality data across all available months.', S['body']))
        E.append(Spacer(1, 4))
        for i, item in enumerate(ai['insights'], 1):
            E.append(Paragraph(f"{i}. {item['title']}", S['ins']))
            E.append(Paragraph(item['body'], S['body']))
            E.append(Spacer(1, 4.5))
    else:
        ins = _rule_insights(D, A, P, PL, prev, wc, gc)
        for i, (t, b) in enumerate(ins, 1):
            E.append(Paragraph(f'{i}. {t}', S['ins'])); E.append(Paragraph(b, S['body'])); E.append(Spacer(1, 4.5))
    E.append(PageBreak())

    # ===== 9 RECOMMENDATIONS =====
    sec('9', 'Priority Recommendations')
    if ai and ai.get('recommendations'):
        order = {'URGENT': 0, 'HIGH': 1, 'MEDIUM': 2, 'MONITOR': 3, 'POSITIVE': 4}
        rr = sorted(ai['recommendations'], key=lambda r: order.get(str(r.get('priority', '')).upper(), 9))
        recs = [['Priority', 'Recommendation']] + [[str(r.get('priority', '—')).upper(), r['text']] for r in rr]
    else:
        recs = _rule_recs(D, A, P, PL, prev, wc, gc, ws_, wd, gs, gd, bl_drift, tot_excess2, pct, prev_name)
    tbl(recs, colw=[22 * mm, 160 * mm], fs=7.0, align='LEFT')
    E.append(Paragraph(f'Auto-generated {datetime.now():%Y-%m-%d %H:%M} by SigmaCement_bot. All figures computed '
                       f'deterministically from the Summary sheet (house rules and naming maps applied); '
                       f'Sections 8-9 written by the analytical layer from those figures'
                       + ('.' if ai else ' (rule-based fallback — AI layer unavailable).'), S['small']))
    doc.build(E)

# ------------------------------------------------------------- ENTRYPOINT ---
def prepare(file_bytes, year, month, prev=None, prev2=None, elec_cost=None):
    """Extract + analyze only. Returns (D, A, metrics, detail) for the AI layer."""
    D = extract(file_bytes)
    A = analyze(D, year, month, prev=prev, prev2=prev2, elec_cost=elec_cost)
    return D, A, metrics_json(D, A, year, month), detail_pack(D, A)


def render(D, A, year, month, ai=None, out_dir=None):
    """Build charts + PDF. `ai` = dict from insight_engine (or None -> rule-based)."""
    tmp = out_dir or tempfile.mkdtemp(prefix='sigma_report_')
    ch = os.path.join(tmp, 'charts'); os.makedirs(ch, exist_ok=True)
    make_charts(D, A, ch, year, month)
    out = os.path.join(tmp, f'production_report_{year}-{month:02d}.pdf')
    build_pdf(D, A, ch, out, year, month, ai=ai)
    return out


def generate_report(file_bytes, year, month, prev=None, prev2=None, elec_cost=None,
                    out_dir=None, ai=None):
    """Returns (pdf_path, metrics_dict). prev/prev2 = metrics dicts of earlier months."""
    D, A, metrics, _detail = prepare(file_bytes, year, month, prev, prev2, elec_cost)
    return render(D, A, year, month, ai=ai, out_dir=out_dir), metrics

def quick_alerts_text(m):
    """Format a metrics dict as a compact Telegram alert message (no AI)."""
    L = [f"🔔 *Alerts — {calendar.month_name[m['month']]} {m['year']}*", '']
    ws_, wd, wc = m['white_pool']; gs, gd, gc = m['grey_pool']
    if wc and wc < WHITE_FLOOR: L.append(f"🔴 White clinker pool: *{wc:.1f} mo* (< {WHITE_FLOOR:.0f} mo floor) — order ~{WHITE_FLOOR*wd-ws_:,.0f} t")
    if gc and gc < GREY_FLOOR + 0.5: L.append(f"🟠 Grey clinker pool: *{gc:.1f} mo* (floor {GREY_FLOOR:.0f} mo)")
    if m['silofull_h'] > 10: L.append(f"🔴 Silos-full stops: *{m['silofull_h']:.1f} h* — dispatch bottleneck")
    if m['incident_h'] > 20: L.append(f"🟠 Incident downtime: *{m['incident_h']:.1f} h*")
    for p, al in m['alerts'].items():
        if al.get('bl_low'): L.append(f"🟠 {p}: Blaine below min on days {[d for d, _ in al['bl_low']]}")
    tot_ex = sum(rd['excess_t'] for rd in m['recipe_dev'].values())
    if tot_ex > 5: L.append(f"🟠 Excess clinker: ~{tot_ex:.0f} t on restart transients")
    if m.get('missing_days'): L.append(f"🟠 Missing daily sheets: {m['missing_days']}")
    if len(L) == 2: L.append('✅ No active alerts — clean month.')
    return '\n'.join(L)

def detail_pack(D, A):
    """Day-level evidence for the AI analytical layer (numbers only, no prose)."""
    daily_tot = {d: round(sum((r[p]['prod'] or 0) for p in r), 1) for d, r in D['daily'].items()}
    dp = {}
    for p in LIMITS:
        rows = {d: r[p] for d, r in D['daily'].items() if p in r}
        if not rows: continue
        dp[p] = {str(d): {k: (round(v, 3) if isinstance(v, float) else v)
                          for k, v in rows[d].items() if v is not None}
                 for d in sorted(rows)}
    return {
        'daily_totals': daily_tot,
        'fridays': A['fridays'],
        'zero_days': D['zero_days'],
        'missing_daily_sheets': D['missing_days'],
        'stoppage_log': [{'day': s['day'], 'hours': s['hours'], 'dept': s['dept'],
                          'reason': s['reason'], 'category': categorize(s)} for s in D['stoppages']],
        'stoppage_by_category': A['stop_cats'],
        'silofull_events': A['silofull_events'],
        'fan_events': A['fan_events'],
        'pi_planned_by_day': D['pi'].get('planned', {}),
        'pi_incident_by_day': D['pi'].get('incident', {}),
        'daily_products': dp,
        'stock_movements': A['stock'],
        'silo_levels': A.get('silo_recon'),
        'packing_by_product': A.get('packing'),
        'm50_bulk_outflow_t': A.get('m50_bulk'),
        'packed_products_unaccounted_t': A.get('packed_gap'),
        'recipe_norms_pct': A['recipe_norm'],
        'recipe_deviations': A['recipe_dev'],
        'product_thresholds': {p: {'tph_min': v[0], 'blaine_min': v[1], 'blaine_max': v[2]}
                               for p, v in LIMITS.items()},
    }


def compute_metrics(file_bytes, year, month, prev=None, prev2=None, elec_cost=None):
    """Lightweight: extraction + analytics only (no charts/PDF). Used at upload time."""
    D = extract(file_bytes)
    A = analyze(D, year, month, prev=prev, prev2=prev2, elec_cost=elec_cost)
    return metrics_json(D, A, year, month)

def detect_content_month(file_bytes):
    """Read the workbook's internal dates (PI/Power/data headers) -> (year, month) or None.
    Used to verify the filename-derived month key against actual content."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        for sn in ('PI', 'Power', 'data', 'DATA', 'Stock'):
            if sn in wb.sheetnames:
                row = next(wb[sn].iter_rows(max_row=1, values_only=True))
                dates = [v for v in row if isinstance(v, datetime)]
                if dates:
                    wb.close()
                    from collections import Counter
                    ym = Counter((d.year, d.month) for d in dates).most_common(1)[0][0]
                    return ym
        # fallback: first daily sheet A1
        import re as _re
        for s in wb.sheetnames:
            if _re.match(r'(Daily report\s+\d+|\d+\s+\w+)', s.strip(), _re.IGNORECASE):
                a1 = next(wb[s].iter_rows(max_row=1, values_only=True))[0]
                if isinstance(a1, datetime):
                    wb.close(); return (a1.year, a1.month)
        wb.close()
    except Exception:
        pass
    return None
